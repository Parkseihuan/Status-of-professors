import openpyxl
import json

# Read the Excel file
file_path = '2025.10.01-보직자현황.xlsx'

try:
    wb = openpyxl.load_workbook(file_path)
    print(f"Sheet names: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*60}")
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}\n")
        
        # Print first 20 rows to understand structure
        print("First 20 rows:")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 20:
                break
            print(f"Row {i}: {row}")
        
        # Also save all data to JSON for web use
        all_data = []
        for row in ws.iter_rows(values_only=True):
            all_data.append(list(row))
        
        with open(f'{sheet_name}_data.json', 'w', encoding='utf-8') as f:
            json.dump(all_data, f, ensure_ascii=False, indent=2)
        print(f"\nData saved to {sheet_name}_data.json")
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
